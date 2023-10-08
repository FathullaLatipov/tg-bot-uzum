import telebot
import buttons
import database
import openpyxl
import datetime
from dotenv import load_dotenv
import os
from telebot.types import ReplyKeyboardRemove

load_dotenv()

bot = telebot.TeleBot(os.getenv('KEY'))

# Путь к файлу Excel
excel_file_path = 'user_data.xlsx'

# Путь к файлу для хранения даты последней отправки отчета
last_report_sent_file = 'last_report_sent.txt'


@bot.message_handler(commands=['start', 'promokod'])
def start(message):
    user_id = message.from_user.id
    bot.send_message(user_id, "Выберите язык / Tilni tanlang", reply_markup=buttons.language_kb())
    bot.register_next_step_handler(message, poluchit_promokod)


def poluchit_promokod(message):
    user_id = message.from_user.id
    if message.text == "Русский язык":
        bot.send_message(user_id, 'Чтобы получить промокод, нажмите "Получить промокод"',
                         reply_markup=buttons.promokod())
        bot.register_next_step_handler(message, register_user)
    elif message.text == "O'zbek tili":
        bot.send_message(user_id, 'Promo-kodni olish uchun “Promo-kodni olish” tugmasini bosing.',
                         reply_markup=buttons.promokod_uz())
        bot.register_next_step_handler(message, register_user_uz)
    else:
        bot.register_next_step_handler(message, start)


def register_user(message):
    user_id = message.from_user.id
    bot.send_message(user_id, "Отправьте фото вашего заказа", reply_markup=ReplyKeyboardRemove())
    bot.register_next_step_handler(message, send_image1)


def send_image1(message):
    user_id = message.from_user.id
    if message.photo:
        photo1 = message.photo[-1].file_id
        bot.send_message(user_id, "Отправьте фото отзыва с оценка 5 звёзд на товар")
        bot.register_next_step_handler(message, send_image2, photo1)
    else:
        bot.send_message(user_id, "Отправьте фотографию, а не файл или текст")
        bot.register_next_step_handler(message, send_image1)


def send_image2(message, photo1):
    user_id = message.from_user.id
    name = message.from_user.username

    save_user_to_excel(user_id, name)
    if message.photo:
        photo2 = message.photo[-1].file_id
        photos = [telebot.types.InputMediaPhoto(photo1), telebot.types.InputMediaPhoto(photo2)]
        bot.send_message(user_id, "Успешно! Мы пришлём промокод  в течении трёх рабочих дней",
                         reply_markup=ReplyKeyboardRemove())
        bot.send_media_group(-4013840171, media=photos)
        bot.send_message(-4013840171, f"<b>Имя</b>: {name}\n"
                                      f"<b>ID</b>: {user_id}\n"
                                      f"<b>Язык</b>: русский",
                         parse_mode="html", reply_markup=buttons.promo_call(user_id))
    else:
        bot.send_message(user_id, "Отправьте фотографию, а не файл или текст")
        bot.register_next_step_handler(message, send_image2_uz, photo1)


def register_user_uz(message):
    user_id = message.from_user.id
    bot.send_message(user_id, "Buyurtmangizning fotosuratini yuboring", reply_markup=ReplyKeyboardRemove())
    bot.register_next_step_handler(message, send_image1_uz)


def send_image1_uz(message):
    user_id = message.from_user.id
    if message.photo:
        photo1 = message.photo[-1].file_id
        bot.send_message(user_id, "Mahsulot uchun 5 yulduzli reyting bilan fotosurat sharhini yuboring")
        bot.register_next_step_handler(message, send_image2_uz, photo1)
    else:
        bot.send_message(user_id, "Fayl yoki matn emas, faqat fotosurat yuboring")
        bot.register_next_step_handler(message, send_image1_uz)


def send_image2_uz(message, photo1):
    user_id = message.from_user.id
    name = message.from_user.username

    save_user_to_excel(user_id, name)
    if message.photo:
        photo2 = message.photo[-1].file_id
        photos = [telebot.types.InputMediaPhoto(photo1), telebot.types.InputMediaPhoto(photo2)]
        bot.send_message(user_id, "Muvaffaqiyatli! Biz uch ish kuni ichida reklama kodini yuboramiz",
                         reply_markup=ReplyKeyboardRemove())
        bot.send_media_group(-4013840171, media=photos)
        bot.send_message(-4013840171, f"<b>Имя</b>: {name}\n"
                                      f"<b>ID</b>: {user_id}\n"
                                      f"<b>Язык</b>: узбекский",
                         parse_mode="html", reply_markup=buttons.promo_call(user_id))
    else:
        bot.send_message(user_id, "Fayl yoki matn emas, faqat fotosurat yuboring")
        bot.register_next_step_handler(message, send_image2_uz, photo1)


def send_promo(message, user_id, admin_id):
    if message.from_user.id == admin_id:
        if message.text == "❌Отмена":
            bot.send_message(-4013840171, f"{user_id} не получил промокод ❌")
        else:
            try:
                bot.send_message(user_id, message.text)
                bot.send_message(-4013840171, f"{user_id} получил промокод✅")
            except:
                bot.send_message(-4013840171, f"{user_id} не получил промокод ❌")


@bot.callback_query_handler(func=lambda call: True)
def promo(call):
    user_id = int(call.data)
    admin_id = call.from_user.id
    bot.send_message(-4013840171,
                     f"Отправьте сообщение для юзера с айди: {user_id}. Для отмены нажмите на кнопку отмены",
                     reply_markup=buttons.cancel_kb())
    bot.register_next_step_handler(call.message, send_promo, user_id, admin_id)


def save_user_to_excel(user_id, username):
    if not os.path.exists(excel_file_path):
        # Создаем новый файл Excel и добавляем заголовки
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['User ID', 'Username', ])
    else:
        # Открываем существующий файл Excel
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active

    # Добавляем данные пользователя в Excel
    ws.append([user_id, username])

    # Получаем максимальную длину имени пользователя
    max_username_length = max(len(row[1].value) for row in ws.iter_rows(min_row=2, max_row=ws.max_row))

    # Устанавливаем ширину столбца "Username" на основе максимальной длины имени
    ws.column_dimensions['B'].width = max_username_length + 2  # Добавляем немного места для отступа

    # Сохраняем файл
    wb.save(excel_file_path)


def send_info_to_group(user_id, username, phone_number, photo):
    bot.send_photo(-4013840171, photo=photo, caption=f"<b>Имя 👤</b>: {username}\n"
                                                     f"<b>ID 🆔</b>: {user_id}\n"
                                                     f"<b>Номер телефона 📞</b>: {phone_number}",
                   parse_mode="html")


def send_excel_report_to_group():
    if os.path.exists(excel_file_path):
        today = datetime.date.today()
        last_modified = datetime.date.fromtimestamp(os.path.getmtime(excel_file_path))

        # Проверяем, был ли файл изменен в последние 30 дней
        if (today - last_modified).days <= 30:
            # Проверяем, отправляли ли отчет сегодня
            if not os.path.exists(last_report_sent_file):
                bot.send_document(-4013840171, open(excel_file_path, 'rb'))
                save_last_report_sent_date(today)


def get_last_report_sent_date():
    if os.path.exists(last_report_sent_file):
        with open(last_report_sent_file, 'r') as file:
            last_report_date = datetime.datetime.strptime(file.read(), '%Y-%m-%d').date()
            return last_report_date
    else:
        return None


def save_last_report_sent_date(date):
    with open(last_report_sent_file, 'w') as file:
        file.write(date.strftime('%Y-%m-%d'))


# Запуск рассылки отчета каждые 30 дней
send_excel_report_to_group()

# Запуск бота
bot.infinity_polling()
